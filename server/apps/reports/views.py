from django.http import HttpResponse

from drf_spectacular.utils import extend_schema, OpenApiParameter

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side,
)

from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView

from rest_framework.decorators import (
    api_view,
    authentication_classes,
    permission_classes,
)

from django.shortcuts import get_object_or_404

from apps.users.infrastructure.authentication import (
    CookieJWTAuthentication,
)

from apps.users.infrastructure.permissions import (
    IsSysAdminOrCoordinator,
)
from apps.files.services.get_files_report import get_files_report

from .models import Notification
from .serializers import NotificationSerializer

from apps.academics.services.course_service import (
    get_courses_data,
    get_top_courses_by_score,
)

from apps.academics.services.teacher_service import (
    get_teachers_stats,
)

from apps.users.service.user_service import (
    get_users_data,
)

from apps.historical.services.courses_historical import (
    get_courses_evolution_data_service,
)
from openpyxl.utils import get_column_letter

def auto_adjust_columns(ws):

    for col_idx, column_cells in enumerate(ws.columns, 1):

        max_length = 0
        column_letter = get_column_letter(col_idx)

        for cell in column_cells:

            if cell.value is None:
                continue

            cell_length = len(str(cell.value))

            if cell_length > max_length:
                max_length = cell_length

        adjusted_width = max_length + 4

        ws.column_dimensions[column_letter].width = min(adjusted_width, 60)

def aplicar_estilos(ws):

    title_fill = PatternFill(
        start_color="111827",
        end_color="111827",
        fill_type="solid"
    )

    title_font = Font(
        bold=True,
        size=16,
        color="FFFFFF"
    )

    header_fill = PatternFill(
        start_color="FACC15",
        end_color="FACC15",
        fill_type="solid"
    )

    header_font = Font(
        bold=True,
        color="000000"
    )

    center = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )

    border = Border(
        left=Side(style="thin", color="2A2A2A"),
        right=Side(style="thin", color="2A2A2A"),
        top=Side(style="thin", color="2A2A2A"),
        bottom=Side(style="thin", color="2A2A2A"),
    )

    for cell in ws[1]:
        cell.fill = title_fill
        cell.font = title_font
        cell.alignment = center

    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for row in ws.iter_rows(min_row=3):
        for cell in row:
            cell.alignment = center
            cell.border = border

    from openpyxl.utils import get_column_letter

    for col_idx in range(1, ws.max_column + 1):

        col_letter = get_column_letter(col_idx)

        max_length = 0

        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):

            cell = row[0]

            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = min(max_length + 5, 40)

def get_faculty_id(request):
    return getattr(
        request.user,
        "faculty_id_id",
        None
    )
@api_view(["GET"])
@authentication_classes([CookieJWTAuthentication])
@permission_classes([IsSysAdminOrCoordinator])
def reporte_docentes(request):

    faculty_id = get_faculty_id(request)

    if not faculty_id:
        return HttpResponse(status=403)

    data = get_teachers_stats(faculty_id)

    wb = Workbook()
    ws = wb.active

    ws.title = "Docentes"

    ws.merge_cells("A1:F1")

    title = ws["A1"]
    title.value = "REPORTE DE DOCENTES"
    title.font = Font(bold=True, size=16, color="FFFFFF")
    title.fill = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
    title.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 30

    ws.append([
        "Docente",
        "Cursos",
        "Promedio",
        "Tendencia (%)",
        "Evaluaciones",
        "Recomendado (%)"
    ])

    for t in data:
        ws.append([
            t["teacher_name"],
            ", ".join(t["cursos_impartidos"]),
            t["promedio_general"],
            t["tendencia_mejora"],
            t["evaluaciones_total"],
            t["recomendado_vs_otros"],
        ])

    aplicar_estilos(ws)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        "attachment; filename=reporte_docentes.xlsx"
    )

    wb.save(response)

    return response

@api_view(["GET"])
@authentication_classes([CookieJWTAuthentication])
@permission_classes([IsSysAdminOrCoordinator])
def reporte_cursos(request):

    faculty_id = get_faculty_id(request)

    if not faculty_id:
        return HttpResponse(status=403)

    data = get_courses_data(faculty_id)

    wb = Workbook()
    ws = wb.active

    ws.title = "Cursos"

    ws.merge_cells("A1:E1")

    title = ws["A1"]
    title.value = "REPORTE DE CURSOS"
    title.font = Font(bold=True, size=16, color="FFFFFF")
    title.fill = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
    title.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 30

    ws.append([
        "Código",
        "Nombre",
        "Créditos",
        "Score",
        "Trend (%)"
    ])

    for c in data:
        ws.append([
            c["code"],
            c["name"],
            c["credits"],
            c["score"],
            c["trend"],
        ])

    aplicar_estilos(ws)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        "attachment; filename=reporte_cursos.xlsx"
    )

    wb.save(response)

    return response

@api_view(["GET"])
@authentication_classes([CookieJWTAuthentication])
@permission_classes([IsSysAdminOrCoordinator])
def reporte_usuarios(request):

    faculty_id = get_faculty_id(request)

    if not faculty_id:
        return HttpResponse(status=403)

    data = sorted(
        get_users_data(faculty_id),
        key=lambda x: x.get("evaluation_count", 0),
        reverse=True
    )

    wb = Workbook()
    ws = wb.active

    ws.title = "Usuarios"

    center = Alignment(
        horizontal="center",
        vertical="center"
    )

    border = Border(
        left=Side(style="thin", color="2A2A2A"),
        right=Side(style="thin", color="2A2A2A"),
        top=Side(style="thin", color="2A2A2A"),
        bottom=Side(style="thin", color="2A2A2A"),
    )

    ws.merge_cells("A1:G1")

    title = ws["A1"]

    title.value = "REPORTE DE USUARIOS"

    title.font = Font(
        bold=True,
        size=16,
        color="FFFFFF"
    )

    title.fill = PatternFill(
        start_color="111827",
        end_color="111827",
        fill_type="solid"
    )

    title.alignment = center

    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:G2")

    subtitle = ws["A2"]

    subtitle.value = (
        "Usuarios subrayados = usuarios inactivos"
    )

    subtitle.font = Font(
        bold=True,
        italic=True,
        color="B91C1C"
    )

    subtitle.alignment = center

    ws.row_dimensions[2].height = 22

    ws.append([
        "Ranking",
        "Usuario",
        "Correo",
        "Rol",
        "Facultad",
        "Evaluaciones",
        "Estado"
    ])

    header_fill = PatternFill(
        start_color="FACC15",
        end_color="FACC15",
        fill_type="solid"
    )

    header_font = Font(
        bold=True,
        color="000000"
    )

    for cell in ws[3]:

        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for i, u in enumerate(data, start=1):

        ws.append([
            i,
            u["username"],
            u["email"],
            u["role"],
            u["faculty"],
            u.get("evaluation_count", 0),
            "Activo" if u["is_active"] else "Inactivo",
        ])

        current_row = ws.max_row

        for cell in ws[current_row]:

            cell.alignment = center
            cell.border = border

        if not u["is_active"]:

            for cell in ws[current_row]:

                cell.font = Font(
                    bold=True,
                    underline="single",
                    color="000000"
                )

    auto_adjust_columns(ws)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        'attachment; filename="reporte_usuarios.xlsx"'
    )

    wb.save(response)

    return response

@api_view(["GET"])
@authentication_classes([CookieJWTAuthentication])
@permission_classes([IsSysAdminOrCoordinator])
def reporte_general(request):

    faculty_id = get_faculty_id(request)

    if not faculty_id:
        return HttpResponse(status=403)

    teachers = get_teachers_stats(faculty_id)
    courses = get_courses_data(faculty_id)
    users = get_users_data()

    wb = Workbook()

    # DOCENTES
    ws1 = wb.active

    ws1.title = "Docentes"

    ws1.append([
        "Docente",
        "Cursos",
        "Promedio",
        "Tendencia (%)"
    ])

    for t in teachers:
        ws1.append([
            t["teacher_name"],
            ", ".join(t["cursos_impartidos"]),
            t["promedio_general"],
            t["tendencia_mejora"],
        ])

    aplicar_estilos(ws1)

    # CURSOS
    ws2 = wb.create_sheet("Cursos")

    ws2.append([
        "Código",
        "Nombre",
        "Créditos",
        "Score",
        "Trend (%)"
    ])

    for c in courses:
        ws2.append([
            c["code"],
            c["name"],
            c["credits"],
            c["score"],
            c["trend"],
        ])

    aplicar_estilos(ws2)

    # USUARIOS
    ws3 = wb.create_sheet("Usuarios")

    ws3.append([
        "Usuario",
        "Correo",
        "Rol",
        "Facultad",
        "Evaluaciones",
        "Estado"
    ])

    for u in users:
        ws3.append([
            u["username"],
            u["email"],
            u["role"],
            u["faculty"],
            u.get("evaluation_count", 0),
            "Activo" if u["is_active"] else "Inactivo",
        ])

    aplicar_estilos(ws3)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        "attachment; filename=reporte_general.xlsx"
    )

    wb.save(response)

    return response


class NotificationListCreateView(APIView):

    authentication_classes = [
        CookieJWTAuthentication
    ]

    permission_classes = [
        IsSysAdminOrCoordinator
    ]

    def get(self, request):

        qs = Notification.objects.all()

        user_id = request.query_params.get("user")

        if user_id:
            qs = qs.filter(user_id=user_id)

        serializer = NotificationSerializer(
            qs,
            many=True
        )

        return Response(serializer.data)

    @extend_schema(
        summary="crear notificación",
        description="Crea una nueva notificación",
        parameters=[
            OpenApiParameter(
                name="notification",
                description="Datos de la notificación a crear",
                required=True,
                type=NotificationSerializer,
            )
        ]
    )
    def post(self, request):

        serializer = NotificationSerializer(
            data=request.data
        )

        if serializer.is_valid():
            serializer.save()

            return Response(
                serializer.data,
                status=status.HTTP_201_CREATED
            )

        return Response(
            serializer.errors,
            status=status.HTTP_400_BAD_REQUEST
        )


class NotificationDetailView(APIView):


    authentication_classes = [
        CookieJWTAuthentication
    ]

    permission_classes = [
        IsSysAdminOrCoordinator
    ]

    def get(self, request, pk):

        notification = get_object_or_404(
            Notification,
            pk=pk
        )

        serializer = NotificationSerializer(
            notification
        )

        return Response(serializer.data)

    def patch(self, request, pk):

        notification = get_object_or_404(
            Notification,
            pk=pk
        )

        serializer = NotificationSerializer(
            notification,
            data=request.data,
            partial=True
        )

        if serializer.is_valid():
            serializer.save()

            return Response(serializer.data)

        return Response(
            serializer.errors,
            status=status.HTTP_400_BAD_REQUEST
        )

    def delete(self, request, pk):

        notification = get_object_or_404(
            Notification,
            pk=pk
        )

        notification.delete()

        return Response(
            status=status.HTTP_204_NO_CONTENT
        )


@api_view(["GET"])
@authentication_classes([CookieJWTAuthentication])
@permission_classes([IsSysAdminOrCoordinator])
def reporte_top_cursos(request):

    faculty_id = get_faculty_id(request)

    if not faculty_id:
        return HttpResponse(status=403)

    semester_id = request.GET.get("semester")

    if not semester_id:
        return HttpResponse(
            "El parámetro semester es requerido",
            status=400
        )

    top4 = get_top_courses_by_score(
        faculty_id,
        semester_id
    )

    wb = Workbook()

    ws = wb.active
    ws.title = "Top 4 Cursos"

    ws.merge_cells("A1:C1")

    title_cell = ws["A1"]
    title_cell.value = "TOP 4 CURSOS CON MEJOR PUNTEO"

    title_cell.font = Font(
        bold=True,
        size=16,
        color="FFFFFF"
    )

    title_cell.fill = PatternFill(
        start_color="111827",
        end_color="111827",
        fill_type="solid"
    )

    title_cell.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    ws.row_dimensions[1].height = 30

    headers = [
        "Ranking",
        "Curso",
        "Punteo"
    ]

    ws.append(headers)

    header_fill = PatternFill(
        start_color="FACC15",
        end_color="FACC15",
        fill_type="solid"
    )

    header_font = Font(
        bold=True,
        color="000000"
    )

    center = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )

    border = Border(
        left=Side(style="thin", color="2A2A2A"),
        right=Side(style="thin", color="2A2A2A"),
        top=Side(style="thin", color="2A2A2A"),
        bottom=Side(style="thin", color="2A2A2A"),
    )

    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for index, course in enumerate(top4, start=1):

        ws.append([
            index,
            course["course_name"],
            course["punteo"],
        ])

    for row in ws.iter_rows(min_row=3):

        for cell in row:
            cell.border = border
            cell.alignment = center

        row[0].font = Font(
            bold=True,
            color="000000"
        )

        row[2].font = Font(
            bold=True,
            color="000000"
        )

    widths = {
        "A": 15,
        "B": 50,
        "C": 20,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = (
        "attachment; filename=top_4_cursos.xlsx"
    )

    wb.save(response)

    return response



@api_view(["GET"])
@authentication_classes([CookieJWTAuthentication])
@permission_classes([IsSysAdminOrCoordinator])
def get_courses_evolution_data(request):

    faculty_id = get_faculty_id(request)

    if faculty_id is None:

        return HttpResponse(status=403)

    data = get_courses_evolution_data_service(faculty_id)

    wb = Workbook()

    ws = wb.active

    ws.title = "Evolución Cursos"



    ws.merge_cells("A1:E1")

    title = ws["A1"]

    title.value = "EVOLUCIÓN HISTÓRICA DE CURSOS"

    title.font = Font(bold=True, size=16, color="FFFFFF")

    title.fill = PatternFill(start_color="111827", end_color="111827", fill_type="solid")

    title.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 30



    headers = ["Curso ID", "Curso", "Año", "Semestre", "Rating"]

    ws.append(headers)

    header_fill = PatternFill(start_color="FACC15", end_color="FACC15", fill_type="solid")

    header_font = Font(bold=True, color="000000")

    center = Alignment(horizontal="center", vertical="center")

    border = Border(

        left=Side(style="thin", color="2A2A2A"),

        right=Side(style="thin", color="2A2A2A"),

        top=Side(style="thin", color="2A2A2A"),

        bottom=Side(style="thin", color="2A2A2A"),

    )

    for cell in ws[2]:

        cell.fill = header_fill

        cell.font = header_font

        cell.alignment = center

        cell.border = border


    for course in data:

        for rating in course.get("semester_ratings", []):

            ws.append([

                course["course_id"],

                course["course_name"],

                rating["semester_year"],

                rating["semester_number"],

                rating["rating"],

            ])



    for row in ws.iter_rows(min_row=3):

        for cell in row:

            cell.border = border

            cell.alignment = center

        row[0].font = Font(bold=True, color="000000")

        row[4].font = Font(bold=True, color="000000")


    ws.column_dimensions["A"].width = 15

    ws.column_dimensions["B"].width = 45

    ws.column_dimensions["C"].width = 15

    ws.column_dimensions["D"].width = 15

    ws.column_dimensions["E"].width = 15


    response = HttpResponse(

        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    )

    response["Content-Disposition"] = 'attachment; filename="evolucion_cursos.xlsx"'

    wb.save(response)

    return response


@api_view(["GET"])
@authentication_classes([CookieJWTAuthentication])
@permission_classes([IsSysAdminOrCoordinator])
def reporte_files_excel(request):

    faculty_id = get_faculty_id(request)

    if not faculty_id:
        return HttpResponse(status=403)

    data = get_files_report(faculty_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Archivos"

    ws.merge_cells("A1:E1")

    title = ws["A1"]

    title.value = "REPORTE DE ARCHIVOS"

    title.font = Font(
        bold=True,
        size=16,
        color="FFFFFF"
    )

    title.fill = PatternFill(
        start_color="111827",
        end_color="111827",
        fill_type="solid"
    )

    title.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:E2")

    summary = ws["A2"]

    summary.value = f"TOTAL DE ARCHIVOS REGISTRADOS: {len(data)}"

    summary.font = Font(
        bold=True,
        color="000000"
    )

    summary.fill = PatternFill(
        start_color="E5E7EB",
        end_color="E5E7EB",
        fill_type="solid"
    )

    summary.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    ws.row_dimensions[2].height = 24

    ws.append([
        "ID",
        "Nombre",
        "Tamaño",
        "Procesado",
        "Fecha"
    ])

    header_fill = PatternFill(
        start_color="FACC15",
        end_color="FACC15",
        fill_type="solid"
    )

    header_font = Font(
        bold=True,
        color="000000"
    )

    for cell in ws[3]:

        cell.fill = header_fill

        cell.font = header_font

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    for f in data:

        ws.append([
            f["id"],
            f["name"],
            f["size"],
            f["processed"],
            f["uploaded_at"],
        ])

    auto_adjust_columns(ws)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        'attachment; filename="reporte_files.xlsx"'
    )

    wb.save(response)

    return response