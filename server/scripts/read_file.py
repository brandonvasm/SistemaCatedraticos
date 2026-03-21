from pathlib import Path
from pprint import pprint

from openpyxl import load_workbook


def normalize_header(header: object, index: int) -> str:
    if header is None:
        return f"column_{index}"

    return str(header).strip().replace("\n", " ")


def is_empty_row(row: tuple) -> bool:
    return all(cell is None for cell in row)


def build_basic_info(workbook, worksheet, headers: list[str], records: list[dict]) -> dict:
    return {
        "sheet_name": worksheet.title,
        "total_sheets": len(workbook.sheetnames),
        "sheet_names": workbook.sheetnames,
        "total_columns": len(headers),
        "total_rows_in_sheet": worksheet.max_row,
        "total_records": len(records),
        "headers": headers,
    }


def read_excel_as_records(
    file_path: str | Path,
    sheet_name: str | None = None,
) -> tuple[dict, list[dict]]:
    path = Path(file_path)

    if not path.exists():
        raise FileNotFoundError(f"No existe el archivo: {path}")

    workbook = load_workbook(filename=path, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active

    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        empty_info = {
            "sheet_name": worksheet.title,
            "total_sheets": len(workbook.sheetnames),
            "sheet_names": workbook.sheetnames,
            "total_columns": 0,
            "total_rows_in_sheet": worksheet.max_row,
            "total_records": 0,
            "headers": [],
        }
        return empty_info, []

    headers = [
        normalize_header(header, index)
        for index, header in enumerate(rows[0], start=1)
    ]

    records: list[dict] = []

    for row in rows[1:]:
        if is_empty_row(row):
            continue

        record = {
            headers[index]: row[index] if index < len(row) else None
            for index in range(len(headers))
        }
        records.append(record)

    basic_info = build_basic_info(workbook, worksheet, headers, records)
    return basic_info, records


def print_basic_info(info: dict) -> None:
    print("\nInformacion del excel")
    print(f"Hoja activa: {info['sheet_name']}")
    print(f"Total de hojas: {info['total_sheets']}")
    print(f"Hojas: {info['sheet_names']}")
    print(f"Total de columnas: {info['total_columns']}")
    print(f"Total de filas en la hoja: {info['total_rows_in_sheet']}")
    print(f"Total de registros(filas del excel sin contar encabezados): {info['total_records']}")
    print("Encabezados:")
    for header in info["headers"]:
        print(f"  - {header}")


def print_records(records: list[dict]) -> None:
    print(f"\n=== MOSTRANDO {len(records)} REGISTROS ===")
    pprint(records, sort_dicts=False)


def main() -> None:
    ## Cambia el path al archivo Excel que deseas leer (se debe de colocar el archivo en la carpeta scripts o ajustar el path a la carpeta que desees)
    file_path = Path("scripts/Control docente.xlsx")

    basic_info, records = read_excel_as_records(file_path=file_path)

    print_basic_info(basic_info)

    if not records:
        print("\nNo se encontraron registros.")
        return

    print_records(records)


main()